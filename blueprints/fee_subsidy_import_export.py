from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from datetime import datetime
import openpyxl
import re
import logging
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from decimal import Decimal
from utils.db import db
from models.fee_subsidy import FeeSubsidy
from models.user import User
from models.room import Room
from models.system_config import SystemConfig
from models.dorm import Dorm  # 新增：导入住宿模型
from utils.log import log_operation  # 导入日志工具
# 导入admin_required装饰器
from blueprints.system_settings import admin_required
import pandas as pd
from utils.excel_date_utils import excel_date_utils

# 创建导入导出专用蓝图
fee_subsidy_import_export_bp = Blueprint('fee_subsidy_import_export', __name__, url_prefix='/fee_subsidy_import_export')

# 优化：通过姓名查询用户ID（增强容错和提示）
def get_user_id_by_name(name):
    if not name:
        return None, "姓名不能为空"
    
    name_str = str(name).strip()
    if not name_str:
        return None, "姓名不能为纯空格"
    
    # 支持模糊查询（提高匹配成功率）
    users = User.query.filter(User.name.ilike(f'%{name_str}%')).all()
    
    if not users:
        return None, f"未找到姓名包含'{name_str}'的用户，请检查姓名是否正确"
    if len(users) > 1:
        user_info = [f"{user.name}（ID:{user.id}）" for user in users]
        return None, f"找到多个匹配用户：{','.join(user_info)}，请使用用户ID导入"
    
    return users[0].id, None

# 验证用户住宿状态与补贴类型的匹配性
def validate_accommodation_status(user_id, fee_type):
    """验证外宿/住宿补贴与用户住宿状态的匹配性"""
    latest_dorm = Dorm.get_user_latest_dorm(user_id)
    has_active_accommodation = False
    
    if latest_dorm:
        if latest_dorm.status == 'active' or not latest_dorm.check_out_date:
            has_active_accommodation = True
    
    if '外宿' in fee_type and has_active_accommodation:
        return False, "用户存在有效住宿记录，无法发放外宿补贴"
    
    if '住宿' in fee_type and not has_active_accommodation:
        return False, "用户无有效住宿记录，无法发放住宿补贴"
    
    return True, "验证通过"


# 1. 批量导出接口（Excel格式）
@fee_subsidy_import_export_bp.route('/export', methods=['GET'])
@login_required
@admin_required
def export_records():
    # 获取筛选参数
    billing_period = request.args.get('billing_period')
    fee_type = request.args.get('fee_type')  # 费用类型参数
    
    # 构建查询，关联用户表和房间表
    query = FeeSubsidy.query.outerjoin(
        User, 
        FeeSubsidy.user_id == User.id
    ).outerjoin(
        Room,
        FeeSubsidy.room_id == Room.id
    )
    
    # 判断是否需要包含禁用的记录
    include_disabled = request.args.get('include_disabled', 'false').lower() == 'true'
    if not include_disabled:
        query = query.filter(FeeSubsidy.is_enabled == True)  # 不包含禁用记录时，只导出启用的记录
    
    # 账期筛选
    if billing_period:
        query = query.filter(FeeSubsidy.billing_period == billing_period)
    
    # 费用类型筛选
    if fee_type:
        query = query.filter(FeeSubsidy.fee_type == fee_type)
    
    records = query.all()
    record_count = len(records)
    
    # 定义类型分类：仅用量型固定，金额型动态判断
    is_usage_type = fee_type == "房间水电按用量减免"  # 用量型固定为这一种
    # 房间相关的金额型（不需要用户信息）
    is_room_based_amount_type = fee_type in ["房间水电按金额减免"]
    # 其他金额型（需要用户信息）
    is_user_based_amount_type = fee_type and not is_usage_type and not is_room_based_amount_type
    
    # 动态生成表头
    basic_headers = [
        'ID', '费用类型', '生效时间', '是否启用', '创建时间', '更新时间',
        '操作人ID', '变更原因', '账期', '账期开始日', '账期结束日'
    ]
    
    # 根据类型添加特定字段
    if is_usage_type:
        # 用量类型：添加用量字段和房间信息，不包含金额和用户信息
        headers = basic_headers + [
            '电费减免量', '水费减免量', '房间ID', '楼栋', '房间号'
        ]
        sheet_title = f"水电用量减免记录_{billing_period or '全部账期'}"
        
    elif is_room_based_amount_type:
        # 房间相关金额型：添加金额和房间信息，不包含用量和用户信息
        headers = basic_headers + [
            '金额', '房间ID', '楼栋','房间号'
        ]
        sheet_title = f"{fee_type}_记录_{billing_period or '全部账期'}"
        
    elif is_user_based_amount_type:
        # 其他金额型：添加金额和用户信息，不包含用量字段
        headers = basic_headers + [
            '金额', '用户ID', '用户姓名',  '部门', '职位'
        ]
        sheet_title = f"{fee_type}_记录_{billing_period or '全部账期'}"
        
    else:
        # 导出全部类型或未指定类型：显示所有字段
        headers = basic_headers + [
            '金额', '电费减免量', '水费减免量',
            '用户ID', '用户姓名', '部门', '职位', 
            '房间ID', '楼栋','房间号'
        ]
        sheet_title = f"费用补贴全记录_{billing_period or '全部账期'}"
    
    # 创建数据列表
    data = []
    
    if not records:
        # 如果没有记录，添加一行空数据来保持表格结构
        empty_row = {header: '' for header in headers}
        empty_row[headers[0]] = '暂无数据'  # 在第一列显示提示
        data.append(empty_row)
    else:
        for record in records:
            # 获取关联信息
            user = User.query.get(record.user_id) if record.user_id else None
            room = Room.query.get(record.room_id) if record.room_id else None
            room_building = room.building if room else ''
            room_number = room.room_number if room else ''
            
            # 创建数据行字典
            row_data = {
                'ID': record.id,
                '费用类型': record.fee_type,
                '生效时间': record.effective_date.strftime('%Y-%m-%d %H:%M:%S') if record.effective_date else '',
                '是否启用': '是' if record.is_enabled else '否',
                '创建时间': record.create_time.strftime('%Y-%m-%d %H:%M:%S') if record.create_time else '',
                '更新时间': record.update_time.strftime('%Y-%m-%d %H:%M:%S') if record.update_time else '',
                '操作人ID': record.operator_id,
                '变更原因': record.change_reason or '',
                '账期': record.billing_period,
                '账期开始日': record.billing_start_date.strftime('%Y-%m-%d') if record.billing_start_date else '',
                '账期结束日': record.billing_end_date.strftime('%Y-%m-%d') if record.billing_end_date else ''
            }
            
            # 记录级别的类型判断（处理全量导出时的混合类型）
            record_is_usage = record.fee_type == "房间水电按用量减免"
            record_is_room_amount = record.fee_type in ["房间水电按金额减免"]
            record_is_user_amount = not record_is_usage and not record_is_room_amount
            
            # 根据类型添加特定数据
            if is_usage_type or record_is_usage:
                # 用量类型数据
                row_data.update({
                    '电费减免量': record.electric_reduction,
                    '水费减免量': record.water_reduction,
                    '房间ID': record.room_id,
                    '楼栋': room_building,
                    '房间号': room_number
                })
            elif is_room_based_amount_type or record_is_room_amount:
                # 房间相关金额型数据
                row_data.update({
                    '金额': record.amount,
                    '房间ID': record.room_id,
                    '楼栋': room_building,
                    '房间号': room_number
                })
            elif is_user_based_amount_type or record_is_user_amount:
                # 其他金额型数据（带用户信息）
                row_data.update({
                    '金额': record.amount,
                    '用户ID': record.user_id,
                    '用户姓名': user.name if user else '',
                    '部门': user.department if user else '',
                    '职位': user.position if user else ''
                })
            else:
                # 全量导出时的默认处理
                row_data.update({
                    '金额': record.amount,
                    '电费减免量': record.electric_reduction,
                    '水费减免量': record.water_reduction,
                    '用户ID': record.user_id,
                    '用户姓名': user.name if user else '',
                    '部门': user.department if user else '',
                    '职位': user.position if user else '',
                    '房间ID': record.room_id,
                    '楼栋': room_building,
                    '房间号': room_number
                })
            
            # 确保所有表头字段都存在于row_data中
            for header in headers:
                if header not in row_data:
                    row_data[header] = ''
            
            data.append(row_data)
    
    # 使用pandas创建DataFrame
    df = pd.DataFrame(data, columns=headers)
    
    # 创建Excel文件
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_title)
        
        # 获取工作表并设置表头样式
        ws = writer.sheets[sheet_title]
        
        # 设置表头样式（加粗）
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)

    # 生成文件名
    type_suffix = fee_type.replace(' ', '') if fee_type else 'all'
    period_suffix = billing_period if billing_period else 'allperiods'
    # 记录导出成功
    log_operation(
            user_id=current_user.id,
            module='feesubsidy',
            operation_type='batch_import_export',
            action=f"成功导出费用补贴记录，共导出{record_count}条数据",
            result="成功"
    )
    # 记录日志
    logging.info(f'成功导出费用补贴记录，共导出{record_count}条数据，操作人ID：{current_user.id}')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'费用补贴数据导出_{type_suffix}_{period_suffix}_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    )

    

# 2. 批量导入接口（添加默认启用状态）
# 批量导入接口（修复*标记导致的字段识别问题）
@fee_subsidy_import_export_bp.route('/import', methods=['POST'])
@login_required
@admin_required
def import_records():
    # 记录导入操作开始
    filename = request.files['file'].filename if 'file' in request.files else '未知文件'
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未上传文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '未选择文件'}), 400
    
    if not file.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'message': '请上传Excel格式文件（.xlsx）'}), 400
    
    transaction_successful = False
    success_count = 0
    error_records = []
    
    try:
        file_content = BytesIO(file.stream.read())
        wb = openpyxl.load_workbook(file_content, data_only=True)
        ws = wb.active
        
        # 读取表头并移除*标记
        headers = []
        for cell in ws[1]:
            header_value = str(cell.value).strip().replace('*', '') if cell.value else ''
            headers.append(header_value)
        
        # 验证基础表头（包含楼栋和房间号）
        required_headers = ['费用类型', '金额', '减免用水量', '减免用电量', '姓名', '楼栋', '房间号', '生效时间']
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            # 记录日志
            logging.error(f'Excel文件缺少必要的列：{",".join(missing_headers)}')
            raise ValueError(f'Excel文件缺少必要的列：{",".join(missing_headers)}')
        
        # 获取系统配置的费用类型
        allowed_types = SystemConfig.get_config_value('ALLOWANCE_TYPES', [])
        if not allowed_types:
            # 记录日志
            logging.error("系统未配置有效的费用类型，请联系管理员")
            raise ValueError("系统未配置有效的费用类型，请联系管理员")
        
        # 类型分类
        water_electric_keyword = "水电"
        water_electric_types = [t for t in allowed_types if water_electric_keyword in t]
        amount_types = [t for t in allowed_types if water_electric_keyword not in t]
        
        # 批量提取所有生效时间值
        effective_date_values = []
        all_rows = []
        
        # 首先收集所有行和生效时间
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # 检测空行
            is_empty_row = all(cell is None or str(cell).strip() == '' for cell in row)
            if not is_empty_row:
                row_data = dict(zip(headers, row))
                effective_date_values.append(row_data.get('生效时间'))
                all_rows.append((row_num, row))
        
        # 批量解析生效时间
        try:
            parsed_effective_dates = excel_date_utils.parse_excel_date(effective_date_values, field_name='生效时间')
        except Exception as e:
            # 记录日志
            logging.error(f"批量解析生效时间失败：{str(e)}")
            return jsonify({
                'success': False,
                'message': f'批量解析生效时间失败：{str(e)}',
                'error_details': []
            }), 500
        
        # 处理数据行
        for idx, (row_num, row) in enumerate(all_rows):
            try:
                # 检测空行
                is_empty_row = all(cell is None or str(cell).strip() == '' for cell in row)
                if is_empty_row:
                    continue
                
                row_data = dict(zip(headers, row))
                fee_type = row_data.get('费用类型')
                
                if not fee_type or str(fee_type).strip() == '':
                    # 记录日志
                    logging.error(f"第{row_num}行费用类型为空（带*的为必填项）")
                    raise ValueError("费用类型不能为空（带*的为必填项）")
                
                fee_type = str(fee_type).strip()
                if fee_type not in allowed_types:
                    # 记录日志
                    logging.error(f"第{row_num}行费用类型不支持：{fee_type}，允许类型：{','.join(allowed_types)}")
                    raise ValueError(f"不支持的费用类型：{fee_type}，允许类型：{','.join(allowed_types)}")
                
                user_id = None
                room_id = None
                is_water_electric = water_electric_keyword in fee_type
                is_amount_type = not is_water_electric
                
                if is_amount_type:
                    # 金额类验证
                    amount = row_data.get('金额')
                    if not amount or not isinstance(amount, (int, float, Decimal)):
                        # 记录日志
                        logging.error(f"第{row_num}行{fee_type}（金额类）金额为空或无效")
                        raise ValueError(f"{fee_type}（金额类）必须填写有效的金额")
                    
                    name = row_data.get('姓名')
                    # 优化：明确提示姓名解析失败的原因
                    user_id, err = get_user_id_by_name(name)
                    if err:
                        # 记录日志
                        logging.error(f"第{row_num}行{fee_type}关联用户失败：{err}")
                        raise ValueError(f"{fee_type}关联用户失败：{err}")
                    
                    if not user_id:
                        # 记录日志
                        logging.error(f"第{row_num}行{fee_type}未获取到有效用户ID，请检查姓名")
                        raise ValueError(f"{fee_type}未获取到有效用户ID，请检查姓名")
                    
                    # 其他验证...
                
                elif is_water_electric:
                    # 水电类验证
                    building = row_data.get('楼栋')
                    room_number = row_data.get('房间号')
                    
                    if not building or str(building).strip() == '':
                        # 记录日志
                        logging.error(f"第{row_num}行{fee_type}楼栋为空（带*的为必填项）")
                        raise ValueError(f"{fee_type}楼栋不能为空（带*的为必填项）")
                    
                    if not room_number or str(room_number).strip() == '':
                        # 记录日志
                        logging.error(f"第{row_num}行{fee_type}房间号为空（带*的为必填项）")
                        raise ValueError(f"{fee_type}房间号不能为空（带*的为必填项）")
                    
                    building = str(building).strip()
                    room_number = str(room_number).strip()
                    
                    # 直接使用楼栋和房间号查询房间ID
                    room = Room.query.filter(
                        Room.building == building,
                        Room.room_number == room_number
                    ).first()
                    
                    if not room:
                        # 记录日志
                        logging.error(f"第{row_num}行{fee_type}未找到房间：楼栋'{building}'，房间号'{room_number}'")
                        raise ValueError(f"{fee_type}未找到房间：楼栋'{building}'，房间号'{room_number}'")
                    
                    room_id = room.id
                    
                    # 其他验证...
                
                # 处理生效时间
                effective_date_obj = parsed_effective_dates[idx]
                if not effective_date_obj:
                    # 记录日志
                    logging.error(f"第{row_num}行生效时间为空或无效")
                    raise ValueError("生效时间不能为空或无效")
                
                effective_date_str = effective_date_obj.strftime('%Y-%m-%d %H:%M:%S')
                
                # 构造数据并保存
                data = {
                    'fee_type': fee_type,
                    'amount': float(row_data['金额']) if row_data.get('金额') else None,
                    'electric_reduction': float(row_data['减免用电量']) if row_data.get('减免用电量') else None,
                    'water_reduction': float(row_data['减免用水量']) if row_data.get('减免用水量') else None,
                    'effective_date': effective_date_str,
                    'operator_id': current_user.id,
                    'change_reason': row_data.get('变更原因') or '批量导入',
                    'user_id': user_id,
                    'room_id': room_id,
                    'is_enabled': True
                }
                
                FeeSubsidy.add_fee(data)
                success_count += 1
                
            except Exception as e:
                # 记录日志
                logging.error(f"第{row_num}行导入失败：{str(e)}")
                error_records.append({
                    'row': row_num,
                    'data': [str(cell) if cell is not None else 'None' for cell in row],
                    'error': str(e)
                })
        
        # 事务处理
        if not error_records:
            db.session.commit()
            transaction_successful = True
        else:
            db.session.rollback()
            success_count = 0
        log_operation(
                user_id=current_user.id,
                module='feesubsidy',
                operation_type='batch_import_export',
                action=f"成功导入费用补贴记录，共处理{success_count + len(error_records)}条，成功{success_count}条",
                result="成功"
            )
        # 记录日志
        logging.info(f'成功导入费用补贴记录，共处理{success_count + len(error_records)}条，成功{success_count}条，失败{len(error_records)}条，操作人ID：{current_user.id}')
        return jsonify({
            'success': transaction_successful,
            'message': f'导入{"成功" if transaction_successful else "失败"}，处理{success_count + len(error_records)}条，成功{success_count}条，失败{len(error_records)}条',
            'error_details': error_records
        })
    
    except Exception as e:
        db.session.rollback()
        # 记录导入异常
        log_operation(
            user_id=current_user.id,
            module='feesubsidy',
            operation_type='batch_import_export',
            action=f"导入费用补贴记录发生错误: {str(e)}",
            result="失败"
        )
        # 记录日志
        logging.error(f'导入费用补贴记录发生错误：{str(e)}，操作人ID：{current_user.id}')
        return jsonify({
            'success': False,
            'message': f'导入过程发生错误: {str(e)}',
            'error_details': error_records
        }), 500
 
    
    
    
# 3. 生成导入模板接口（更新时间格式说明）
@fee_subsidy_import_export_bp.route('/import-template', methods=['GET'])
@login_required
@admin_required
def generate_import_template():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "费用补贴数据"
        
        # 获取系统配置的费用类型
        allowed_types = SystemConfig.get_config_value(
            'ALLOWANCE_TYPES', 
            default=["外宿补贴", "住宿补贴", "房间水电按用量减免", "房间水电按金额减免"]
        )
        
        if not isinstance(allowed_types, list):
            allowed_types = [str(allowed_types)]
        
        # 类型分组
        water_electric_keyword = "水电"
        water_electric_types = [t for t in allowed_types if water_electric_keyword in t]
        amount_types = [t for t in allowed_types if water_electric_keyword not in t]
        
        # 表头（带*标记）
        headers = [
            "费用类型*", "金额*", "减免用电量*", "减免用水量*",
            "姓名*", "楼栋*", "房间号*", "生效时间*", "变更原因"
        ]
        
        # 写入表头样式
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 类型说明
        ws.cell(row=2, column=1, value="系统当前配置的所有费用类型：").font = Font(bold=True)
        ws.cell(row=2, column=2, value=",".join(allowed_types)).font = Font(color="1F4E78")
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=len(headers))
        
        ws.cell(row=3, column=1, value="金额类类型（需填写姓名）：").font = Font(bold=True)
        amount_types_str = ",".join(amount_types) if amount_types else "无"
        ws.cell(row=3, column=2, value=amount_types_str)
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=len(headers))
        
        ws.cell(row=4, column=1, value="水电类类型（需填写房间号）：").font = Font(bold=True)
        water_types_str = ",".join(water_electric_types) if water_electric_types else "无"
        ws.cell(row=4, column=2, value=water_types_str)
        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=len(headers))
        
        # 示例数据
        examples = []
        for amount_type in amount_types:
            examples.append({
                "费用类型*": amount_type,
                "金额*": 500.00 if '补贴' in amount_type else 300.00,
                "减免用电量*": "",
                "减免用水量*": "",
                "姓名*": "张三",
                "楼栋*": "A楼栋",
                "房间号*": "102",
                "生效时间*": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "变更原因": f"{amount_type}（示例）"
            })
        
        for water_type in water_electric_types:
            examples.append({
                "费用类型*": water_type,
                "金额*": "" if '用量' in water_type else 200.00,
                "减免用电量*": 50.0 if '用量' in water_type else "",
                "减免用水量*": 10.0 if '用量' in water_type else "",
                "姓名*": "",
                "楼栋*": "A楼栋",
                "房间号*": "101",
                "生效时间*": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "变更原因": f"{water_type}（示例）"
            })
        
        # 写入示例数据
        for row_idx, example in enumerate(examples, 6):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=example[header])
                cell.fill = PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid")
        
        # 强化说明文本（核心修改：明确要求删除说明和示例行）
        instructions = [
            "===== 重要！导入前必须执行以下操作 =====",
            "1. 删除本文件中所有说明行（包括本行及以上的所有文字说明）",
            "2. 删除所有示例数据行（带浅蓝色背景的行）",
            "3. 仅保留表头行（第一行，带蓝色背景的行）和您要导入的数据行",
            "4. 数据行必须从表头行的下一行（即原第2行）开始填写，中间不能有空行",
            "",
            "===== 填写规则 =====",
            f"1. 费用类型必须严格匹配系统配置：{','.join(allowed_types)}",
            "2. 带*的字段为必填项，根据类型填写：",
            f"   - 金额类（{amount_types_str}）：填写金额、姓名，其他数值字段留空",
            f"   - 水电类（{water_types_str}）：填写楼栋、房间号，按需填写金额或水电减免量",
            "3. 楼栋和房间号为独立必填项，需分别填写，不得合并填写",
            "4. 生效时间格式：推荐YYYY-MM-DD HH:MM:SS（如2023-08-17 14:30:00）",
            "5. 姓名必须唯一，重名用户请使用用户ID导入"
        ]
        
        # 写入说明
        first_instr_row = len(examples) + 8
        for i, instr in enumerate(instructions, first_instr_row):
            ws.cell(row=i, column=1, value=instr).font = Font(size=10)
            if "重要" in instr:
                ws.cell(row=i, column=1).font = Font(size=10, bold=True, color="FF0000")
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(headers))
        
        # 调整列宽
        column_widths = [15, 10, 12, 12, 10, 15, 20, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # 保存模板
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        # 记录模板生成成功
        filename = f'费用补贴导入模板_{datetime.now().strftime("%Y%m%d")}.xlsx'
        log_operation(
            user_id=current_user.id,
            module='feesubsidy',
            operation_type='batch_import_export',
            action=f"成功生成费用补贴导入模板，文件名: {filename}",
            result="成功"
        )
        # 记录日志
        logging.info(f'成功生成费用补贴导入模板，文件名：{filename}，操作人ID：{current_user.id}')
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'费用补贴导入模板_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        logging.error(f"生成导入模板失败: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'生成模板失败: {str(e)}'
        }), 500