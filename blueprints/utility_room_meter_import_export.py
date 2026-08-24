from flask import Blueprint, request, jsonify, make_response
from utils.db import db
from datetime import datetime
import logging
import io
import email.utils
import calendar
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from models.utility_room_meter import UtilityMeterReading
from models.room import Room
from utils.log import log_operation
from sqlalchemy.exc import SQLAlchemyError
from flask_login import current_user, login_required
from urllib.parse import quote
import pandas as pd
from decimal import Decimal  # 确保导入Decimal
from io import BytesIO
# 导入admin_required装饰器
from blueprints.system_settings import admin_required
from utils.excel_date_utils import excel_date_utils


# 创建蓝图
utility_room_meter_import_export_bp = Blueprint('utility_room_meter_import_export', __name__, url_prefix='/utility_room_meter_import_export')


# 模板下载接口（不含抄表人字段）
@utility_room_meter_import_export_bp.route('/template', methods=['GET'])
@login_required
@admin_required
def download_template():
    """下载抄表记录导入模板（不含抄表人字段）"""
    try:
        # 创建Excel工作簿作为模板
        wb = Workbook()
        ws = wb.active
        ws.title = "抄表记录模板"

        # 表头（不含抄表人字段）
        headers = [
            "记录ID（批量更新必填）", "抄表日期时间", "楼栋", "宿舍号",
            "水表本次读数", "水表是否更换", "水表备注",
             "电表本次读数", "电表是否更换", "电表备注",
             "抄表类型"
        ]
        ws.append(headers)

        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 添加示例数据行
        ws.append([
            "",  # 新增记录无需填写ID
            "2024-06-01 10:30:00", "3", "305",
            "567.2", "589.7", "否", "表具正常",
            "1234.0", "1256.8", "否", "表具正常",
            "退宿抄表"
        ])
        
        ws.append([
            "1001",  # 批量更新时填写此ID
            "2024-06-30 10:30:00", "3", "306",
            "590.1", "610.5", "是", "更换新表",
            "1300.2", "1320.3", "否", "表具正常",
            "正常抄表"
        ])

        # 调整列宽
        column_widths = [18, 25, 8, 8, 12, 12, 10, 20, 12, 12, 10, 20, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 处理文件名编码
        filename = "抄表记录导入模板.xlsx"
        encoded_filename = quote(filename)
        
        # 构建响应
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"
        response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        # 记录模板下载日志
        log_operation(
            user_id=current_user.id,
            module="utility",
            operation_type="batch_import_export",
            action=f"下载导入模板,文件名:{filename}",
            result="成功"
        )
        logging.info(f"{current_user.id}下载抄表记录导入模板成功")
        return response

    except Exception as e:
        logging.exception("模板下载过程发生错误")
        error_msg = f"模板下载失败: {str(e)}"
        log_operation(
            user_id=current_user.id,
            module="utility",
            operation_type="batch_import_export",
            action=f"模板下载失败{error_msg}",
            result="失败"
        )
        return jsonify({"success": False, "message": error_msg}), 500
    
# 导出抄表记录为Excel（包含记录ID，不含抄表人信息）
# 导出抄表记录为Excel（包含记录ID，不含抄表人信息）
@utility_room_meter_import_export_bp.route('/export', methods=['GET'])
@login_required
@admin_required
def export_readings():
    """导出抄表记录为Excel文件（支持按账期筛选，包含记录ID，不含抄表人信息）"""
    try:
        # 获取筛选参数
        billing_period = request.args.get('billing_period')

        # 参数验证：只允许billing_period参数
        if not billing_period:
            logging.error("导出抄表记录失败：缺少billing_period参数")
            return jsonify({
                'success': False, 
                'message': '请提供billing_period参数（格式：YYYY-MM）'
            }), 400

        # 构建查询
        query = UtilityMeterReading.query.join(Room)

        # 账期筛选逻辑
        try:
            # 验证billing_period格式
            datetime.strptime(billing_period, '%Y-%m')
            
            # 导入RoomUtilityRecord模型
            from models.utility_room_bill_record import RoomUtilityRecord
            
            # 根据账期查询所有主表记录
            main_records = RoomUtilityRecord.query.filter_by(billing_period=billing_period).all()
            
            if not main_records:
                logging.error(f"导出抄表记录失败：未找到{billing_period}账期的记录")
                return jsonify({
                    'success': False, 
                    'message': f'未找到{billing_period}账期的记录'
                }), 404
            
            # 提取所有主表记录ID
            record_ids = [record.record_id for record in main_records]
            
            # 根据record_id筛选抄表记录
            query = query.filter(UtilityMeterReading.record_id.in_(record_ids))
        except ValueError:
            logging.error(f"导出抄表记录失败：账期格式错误，{billing_period}")
            return jsonify({
                'success': False, 
                'message': '账期格式错误，请使用YYYY-MM格式'
            }), 400

        # 按日期倒序
        readings = query.order_by(UtilityMeterReading.reading_date.desc()).all()
        record_count = len(readings)

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "抄表记录"

        # 表头（移除抄表人相关字段）
        headers = [
            "记录ID（批量更新必填）", "序号", "抄表日期时间", "楼栋", "宿舍号",
            "抄表类型", "水表上次读数", "水表本次读数", "用水量(m³)", "水表是否更换", "水表备注",
            "电表上次读数", "电表本次读数", "用电量(kWh)", "电表是否更换", "电表备注"
        ]
        ws.append(headers)

        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 填充数据（不含抄表人信息）
        for idx, reading in enumerate(readings, 1):
            room = reading.room
            ws.append([
                reading.id,  # 记录ID，用于批量更新时的唯一标识
                idx,
                reading.reading_date.strftime('%Y-%m-%d %H:%M:%S'),
                room.building if room else "",
                room.room_number if room else "",
                "正常抄表" if reading.reading_type == 1 else "退宿抄表" if reading.reading_type == 2 else "",
                reading.water_previous or "",
                reading.water_current or "",
                reading.water_usage or "",
                "是" if reading.water_meter_replaced else "否",
                reading.water_notes or "",
                reading.electric_previous or "",
                reading.electric_current or "",
                reading.electric_usage or "",
                "是" if reading.electric_meter_replaced else "否",
                reading.electric_notes or ""
            ])

        # 调整列宽
        column_widths = [18, 5, 20, 15, 8, 8, 10, 12, 12, 15, 10, 20, 12, 12, 15, 10, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 处理中文文件名编码
        filename = f"抄表记录_{billing_period}_{datetime.now()}.xlsx"
        encoded_filename = quote(filename)
        
        # 构建响应
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"
        response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        # 记录日志
        log_operation(
            user_id=current_user.id,
            module="utility",
            operation_type="batch_import_export",
            action=f"成功导出抄表记录数据，共{record_count}条记录",
            result="成功"
        )
        logging.info(f"成功导出抄表记录数据，共{record_count}条记录")
        return response

    except SQLAlchemyError as e:
        db.session.rollback()
        error_msg = f"数据库错误: {str(e)}"
        log_operation(
            user_id=current_user.id,
            module="utility",
            operation_type="batch_import_export",
            action=f"导出失败{error_msg}",
            result="失败"
        )
        return jsonify({"success": False, "message": error_msg}), 500
    except Exception as e:
        logging.error(f"导出记录发生错误{str(e)}")
        error_msg = f"导出失败: {str(e)}"
        log_operation(
            user_id=current_user.id,
            module="utility",
            operation_type="batch_import_export",
            action=f"导出失败{error_msg}",
            result="失败"
        )
        return jsonify({"success": False, "message": error_msg}), 500


# 导入抄表记录从Excel（新记录导入，无需ID和抄表人）
# 导入抄表记录从Excel（确保触发自动同步逻辑）
@utility_room_meter_import_export_bp.route('/import', methods=['POST'])
@login_required
@admin_required
def import_readings():
    """从Excel导入新的抄表记录（自动同步上次读数）"""
    log_data = {
        "文件名": "",
        "记录总数": 0,
        "成功数": 0,
        "失败数": 0,
        "错误详情": []
    }
    
    try:
        if not current_user.is_authenticated:
            msg = "请先登录系统"
            logging.error(f"导入文件失败{msg}")
            return jsonify({"success": False, "message": msg}), 401

        if 'file' not in request.files:
            msg = "未上传文件"
            log_operation(
                user_id=current_user.id,
                module="utility",
                operation_type="batch_import_export",
                action=f"导入失败{msg}",
                result="失败"
            )
            logging.error(f"导入文件失败{msg}")
            return jsonify({"success": False, "message": msg}), 400

        file = request.files['file']
        log_data["文件名"] = file.filename
        
        if file.filename == '':
            msg = "未选择文件"
            log_operation(
                user_id=current_user.id,
                module="utility",
                operation_type="batch_import_export",
                action=f"导入失败{msg}",
                result="失败"
            )
            logging.error(f"导入文件失败{msg}")
            return jsonify({"success": False, "message": msg}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            msg = "仅支持Excel文件(.xlsx, .xls)"
            log_operation(
                user_id=current_user.id,
                module="utility",
                operation_type="batch_import_export",
                action=f"导入失败{msg}",
                result="失败"
            )
            logging.error(f"导入文件失败{msg}")
            return jsonify({"success": False, "message": msg}), 400

        # 关键修复1：将文件流转换为BytesIO对象
        file_content = file.read()  # 读取文件内容为字节流
        file_bytes = BytesIO(file_content)  # 转换为BytesIO（支持seekable）
        file_bytes.seek(0)  # 重置指针到开头

        # 读取Excel文件（使用转换后的BytesIO对象）
        wb = load_workbook(file_bytes, data_only=True)
        ws = wb.active

        # 验证必要表头（忽略未知表头及内容）
        required_headers = [
            "抄表日期时间", "楼栋", "宿舍号",
            "水表本次读数", "电表本次读数"
        ]
        
        # 获取实际表头
        actual_headers = []
        for cell in ws[1]:
            if cell.value is not None:
                actual_headers.append(str(cell.value).strip())
            else:
                actual_headers.append("")
        
        # 检查所有必要表头是否存在
        missing_headers = [header for header in required_headers if header not in actual_headers]
        if missing_headers:
            msg = f"缺少必要的表头: {', '.join(missing_headers)}"
            log_operation(
                user_id=current_user.id,
                module="utility",
                operation_type="batch_import_export",
                action=f"导入失败{msg}",
                result="失败"
            )
            logging.error(f"导入文件失败{msg}")
            return jsonify({"success": False, "message": msg}), 400
        
        # 记录所有表头索引，用于后续数据读取
        header_indices = {header: idx for idx, header in enumerate(actual_headers)}

        # 处理数据
        success_count = 0
        fail_count = 0
        error_records = []
        # 修复1：准确计算有数据的行数（排除空行）
        total_records = 0
        # 先收集所有行数据和抄表日期时间值
        non_empty_rows = []
        reading_date_values = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # 检查是否有实际数据（至少一个单元格有值）
            if any(cell is not None and str(cell).strip() for cell in row):
                total_records += 1
                non_empty_rows.append((row_num, row))
                # 尝试获取抄表日期时间值
                reading_date_idx = header_indices.get("抄表日期时间")
                if reading_date_idx is not None and reading_date_idx < len(row):
                    reading_date_values.append(row[reading_date_idx])
                else:
                    reading_date_values.append(None)
        log_data["记录总数"] = total_records
        
        # 批量解析抄表日期时间
        try:
            parsed_reading_dates = excel_date_utils.parse_excel_date(reading_date_values, field_name='抄表日期时间')
        except Exception as e:
            # 记录日志
            logging.error(f"批量解析抄表日期时间失败：{str(e)}")
            return jsonify({
                'success': False,
                'message': f'批量解析抄表日期时间失败：{str(e)}',
                'error_details': []
            }), 500
        
        # 处理非空行数据
        for idx, (row_num, row) in enumerate(non_empty_rows):
            try:
                # 动态获取各列数据（忽略未知表头）
                reading_date_value = row[header_indices.get("抄表日期时间")]
                building = row[header_indices.get("楼栋")]
                room_number = row[header_indices.get("宿舍号")]
                water_curr = row[header_indices.get("水表本次读数")]
                water_replaced = row[header_indices.get("水表是否更换")]
                water_notes = row[header_indices.get("水表备注")]
                electric_curr = row[header_indices.get("电表本次读数")]
                electric_replaced = row[header_indices.get("电表是否更换")]
                electric_notes = row[header_indices.get("电表备注")]
                reading_type_str = row[header_indices.get("抄表类型")]
                
                # 解析抄表类型
                reading_type_map = {
                    '正常抄表': 1,
                    '退宿抄表': 2
                }
                
                if reading_type_str is None or str(reading_type_str).strip() == '':
                    reading_type = 1  # 默认正常抄表
                elif str(reading_type_str).strip().isdigit():
                    reading_type = int(str(reading_type_str).strip())
                    # 验证数字范围
                    if reading_type not in [1, 2]:
                        logging.error(f"无效的抄表类型数值: {reading_type}，1(正常抄表)或2(退宿抄表)")
                        raise ValueError(f"无效的抄表类型数值: {reading_type}，1(正常抄表)或2(退宿抄表)")
                else:
                    reading_type = reading_type_map.get(str(reading_type_str).strip(), None)
                    if reading_type is None:
                        logging.error(f"无效的抄表类型: {reading_type_str}，请使用'正常抄表'或'退宿抄表'")
                        raise ValueError(f"无效的抄表类型: {reading_type_str}，请使用'正常抄表'或'退宿抄表'")
                
                # 水表上次读数和电表上次读数由模型自动同步，不需要从Excel读取
                water_prev = None
                electric_prev = None

                # 验证必要字段
                if not all([reading_date_value, building, room_number]):
                    logging.error("抄表日期时间、楼栋、宿舍号为必填项")
                    raise ValueError("抄表日期时间、楼栋、宿舍号为必填项")

                # 使用批量解析的日期
                reading_date = parsed_reading_dates[idx]
                if not reading_date:
                    # 记录日志
                    logging.error(f"第{row_num}行抄表日期时间为空或无效")
                    raise ValueError("抄表日期时间不能为空或无效")

                # 查找房间
                room = Room.query.filter_by(
                    building=str(building).strip(),
                    room_number=str(room_number).strip()
                ).first()
                if not room:
                    logging.error(f"未找到楼栋{building}宿舍{room_number}的有效房间")
                    raise ValueError(f"未找到楼栋{building}宿舍{room_number}的有效房间")

                # 转换数值类型（仅处理本次读数，上次读数由模型自动同步）
                water_current = float(water_curr) if (water_curr is not None and str(water_curr).strip()) else None
                electric_current = float(electric_curr) if (electric_curr is not None and str(electric_curr).strip()) else None

                # 转换布尔值
                water_meter_replaced = str(water_replaced).lower() in ['是', 'true', '1'] if water_replaced else False
                electric_meter_replaced = str(electric_replaced).lower() in ['是', 'true', '1'] if electric_replaced else False

                # 调用模型的create_reading方法（自动同步上次读数）
                # 注意：这里不传递water_prev和electric_prev，由模型自动计算
                new_reading = UtilityMeterReading.create_reading(
                    room_id=room.id,
                    water_current=water_current,
                    electric_current=electric_current,
                    water_meter_replaced=water_meter_replaced,
                    electric_meter_replaced=electric_meter_replaced,
                    reading_date=reading_date,
                    meter_reader_id=current_user.id,  # 默认当前用户
                    water_notes=str(water_notes).strip() if water_notes else None,
                    electric_notes=str(electric_notes).strip() if electric_notes else None,
                    reading_type=reading_type
                )
                
                success_count += 1
                # 提交事务（确保模型计算的字段被保存）
                db.session.commit()

            except Exception as e:
                fail_count += 1
                error = str(e)
                error_records.append({
                    "row": row_num,
                    "error": error,
                    "data": [str(cell) for cell in row]
                })
                logging.error(f"第{row_num}行数据导入失败: {error}")
                if len(log_data["错误详情"]) < 10:
                    log_data["错误详情"].append({"行号": row_num, "错误信息": error})


        log_data["成功数"] = success_count
        log_data["失败数"] = fail_count
        
        log_operation(
            user_id=current_user.id,
            module="utility",
            operation_type="batch_import_export",
            action=f"抄表记录导入完成，共处理{total_records}条，成功{success_count}条，失败{fail_count}条",
            result="成功"
        )
        
        logging.info(f"抄表记录导入完成，成功{success_count}条，失败{fail_count}条，共处理{total_records}条")

        return jsonify({
            "success": True,
            "message": f"导入完成，成功{success_count}条，失败{fail_count}条",
            "total_count": total_records,  # 新增：返回实际处理的总记录数
            "success_count": success_count,
            "fail_count": fail_count,
            "errors": error_records if fail_count > 0 else None
        })

    except SQLAlchemyError as e:
        db.session.rollback()
        error_msg = f"数据库错误: {str(e)}"
        logging.error(error_msg)
        log_operation(
            user_id=current_user.id,
            module="utility",
            operation_type="batch_import_export",
            action=f"导入失败{error_msg}",
            result="失败"
        )
        return jsonify({"success": False, "message": error_msg}), 500
    except Exception as e:
        error_msg = f"导入失败: {str(e)}"
        logging.error(error_msg)
        log_operation(
            user_id=current_user.id,
            module="utility",
            operation_type="batch_import_export",
            action=f"导入失败{error_msg}",
            result="失败"
        )
        return jsonify({"success": False, "message": error_msg}), 500

# 批量更新抄表记录接口（确保触发自动同步逻辑）
@utility_room_meter_import_export_bp.route('/batch_update', methods=['POST'])
@login_required
@admin_required
def batch_update():
    """批量更新抄表记录（自动同步上次读数）"""
    try:
        if not current_user.is_authenticated:
            msg = "请先登录系统"
            logging.error(msg)
            return jsonify({"success": False, "message": msg}), 401

        if 'file' not in request.files:
            logging.error("未上传文件")
            return jsonify({'success': False, 'message': '未上传文件'}), 400

        file = request.files['file']
        if file.filename == '':
            logging.error("未选择文件") 
            return jsonify({'success': False, 'message': '请选择有效的Excel文件'}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            logging.error("文件格式错误")   
            return jsonify({'success': False, 'message': '文件格式错误，请上传Excel文件'}), 400

        # 关键修复2：将文件流转换为BytesIO对象
        file_content = file.read()  # 读取文件内容为字节流
        file_bytes = BytesIO(file_content)  # 转换为BytesIO（支持seekable）
        file_bytes.seek(0)  # 重置指针到开头

        # 读取Excel文件（使用转换后的BytesIO对象）
        df = pd.read_excel(file_bytes)
        
        # 验证必要的表头
        required_columns = ['记录ID（批量更新必填）', '抄表日期时间', '楼栋', '宿舍号']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            error_msg = f'Excel缺少必要的列：{", ".join(missing_columns)}'
            logging.error(error_msg)
            return jsonify({'success': False, 'message': error_msg}), 400

        # 预处理数据
        success_count = 0
        fail_count = 0
        errors = []
        processed_ids = {}  # 缓存已处理的ID，提高效率
        actual_updated = 0  # 记录实际发生变化的记录数

        # 批量提取所有抄表日期时间值
        reading_date_values = []
        for index, row in df.iterrows():
            reading_date_str = str(row['抄表日期时间']).strip() if pd.notna(row['抄表日期时间']) else None
            reading_date_values.append(reading_date_str)
        
        # 批量解析抄表日期时间
        try:
            parsed_reading_dates = excel_date_utils.parse_excel_date(reading_date_values, field_name='抄表日期时间')
        except Exception as e:
            # 记录日志
            logging.error(f"批量解析抄表日期时间失败：{str(e)}")
            return jsonify({
                'success': False,
                'message': f'批量解析抄表日期时间失败：{str(e)}',
                'errors': []
            }), 500

        # 开始数据库事务
        try:
            for index, row in df.iterrows():
                row_num = index + 2
                try:
                    # 提取记录ID
                    record_id_str = str(row['记录ID（批量更新必填）']).strip()
                    if not record_id_str:
                        logging.error("记录ID不能为空（批量更新必须提供）")
                        raise ValueError("记录ID不能为空（批量更新必须提供）")
                    
                    # 验证ID格式
                    try:
                        record_id = int(record_id_str)
                    except ValueError:
                        logging.error(f"记录ID格式错误，当前值: {record_id_str}")
                        raise ValueError(f"记录ID必须是数字，当前值: {record_id_str}")

                    # 提取基础信息
                    building = str(row['楼栋']).strip()
                    room_number = str(row['宿舍号']).strip()
                    reading_date_str = str(row['抄表日期时间']).strip()
                        
                    if not all([building, room_number, reading_date_str]):
                        logging.error(f"第{row_num}行数据缺失楼栋、宿舍号或抄表日期时间")
                        raise ValueError(f'楼栋、宿舍号和抄表日期时间不能为空')

                    # 使用批量解析的日期
                    reading_datetime = parsed_reading_dates[index]
                    if not reading_datetime:
                        logging.error(f"第{row_num}行抄表日期时间为空或无效")
                        raise ValueError("抄表日期时间不能为空或无效")

                    # 通过ID查询记录
                    cache_key = f"id_{record_id}"
                    if cache_key not in processed_ids:
                        reading = UtilityMeterReading.query.filter_by(id=record_id).first()
                        processed_ids[cache_key] = reading

                    reading = processed_ids[cache_key]
                    
                    # 验证记录存在性
                    if not reading:
                        logging.error(f"第{row_num}行记录ID {record_id} 不存在")
                        raise ValueError(f'记录ID {record_id} 不存在')
                    
                    # 验证楼栋宿舍匹配
                    room = Room.query.get(reading.room_id)
                    if not room or room.building != building or room.room_number != room_number:
                        logging.error(f"第{row_num}行记录ID {record_id} 与楼栋{building}宿舍{room_number}不匹配")
                        raise ValueError(f'记录ID {record_id} 与楼栋{building}宿舍{room_number}不匹配')
                    
                    # 验证日期
                    if reading.reading_date.date() != reading_datetime.date():
                        logging.error(f"第{row_num}行记录ID {record_id} 与抄表日期时间不匹配（必须为同一天）")
                        raise ValueError(f'记录ID {record_id} 与抄表日期时间不匹配（必须为同一天）')

                    # 提取更新字段（仅传递需要更新的字段，由模型处理同步逻辑）
                    update_data = {
                        'meter_reader_id': current_user.id,  # 默认当前用户
                        'reading_date': reading_datetime  # 确保时间精确
                    }
                    
                    # 处理水表数据（触发模型自动同步上次读数）
                    if pd.notna(row.get('水表本次读数')):
                        try:
                            update_data['water_current'] = Decimal(str(row['水表本次读数']))
                        except ValueError:
                            logging.error(f"第{row_num}行记录ID {record_id} 水表本次读数格式错误")
                            raise ValueError("水表本次读数必须是数字")
                    
                    # 处理电表数据（触发模型自动同步上次读数）
                    if pd.notna(row.get('电表本次读数')):
                        try:
                            update_data['electric_current'] = Decimal(str(row['电表本次读数']))
                        except ValueError:
                            logging.error(f"第{row_num}行记录ID {record_id} 电表本次读数格式错误")
                            raise ValueError("电表本次读数必须是数字")
                    
                    # 处理更换标记（影响模型同步逻辑）
                    update_data['water_meter_replaced'] = str(row.get('水表是否更换', '')).lower() in ['是', 'true', '1']
                    update_data['electric_meter_replaced'] = str(row.get('电表是否更换', '')).lower() in ['是', 'true', '1']
                    
                    # 处理备注
                    water_notes = str(row.get('水表备注', '')).strip() or None
                    if water_notes is not None:
                        update_data['water_notes'] = water_notes
                    
                    electric_notes = str(row.get('电表备注', '')).strip() or None
                    if electric_notes is not None:
                        update_data['electric_notes'] = electric_notes
                    
                    # 处理抄表类型
                    if '抄表类型' in df.columns and pd.notna(row.get('抄表类型')):
                        reading_type_str = str(row.get('抄表类型')).strip()
                        reading_type_map = {
                            '正常抄表': 1,
                            '退宿抄表': 2
                        }
                        
                        if reading_type_str.isdigit():
                            reading_type = int(reading_type_str)
                            # 验证数字范围
                            if reading_type not in [1, 2]:
                                logging.error(f"第{row_num}行记录ID {record_id} 无效的抄表类型数值: {reading_type}，1(正常抄表)或2(退宿抄表)")
                                raise ValueError(f"无效的抄表类型数值: {reading_type}，1(正常抄表)或2(退宿抄表)")
                        else:
                            reading_type = reading_type_map.get(reading_type_str, None)
                            if reading_type is None:
                                logging.error(f"第{row_num}行记录ID {record_id} 无效的抄表类型: {reading_type_str}，请使用'正常抄表'或'退宿抄表'")
                                raise ValueError(f"无效的抄表类型: {reading_type_str}，请使用'正常抄表'或'退宿抄表'")
                        
                        update_data['reading_type'] = reading_type

                    # 调用模型的update方法（自动同步上次读数和用量）
                    reading.update(** update_data)
                    actual_updated += 1
                    success_count += 1

                except Exception as e:
                    fail_count += 1
                    error_msg = str(e)
                    errors.append({
                        'row': row_num,
                        'record_id': record_id_str if 'record_id_str' in locals() else '',
                        'error': error_msg
                    })
                    logging.warning(f'批量更新行{row_num}失败: {error_msg}')
                    log_operation(
                        user_id=current_user.id,
                        module="utility",
                        operation_type="batch_import_export",
                        action=f'批量更新行{row_num}失败: {error_msg}',
                        result="失败"
                    )
                

            # 提交事务（确保模型计算的字段被保存）
            db.session.commit()

            log_operation(
                user_id=current_user.id,
                module="utility",
                operation_type="batch_import_export",
                action=f'抄表记录批量处理完成，成功{success_count}条，失败{fail_count}条，实际更新{actual_updated}条',
                result="成功"
            )
            logging.info(f'抄表记录批量处理完成，成功{success_count}条，失败{fail_count}条，实际更新{actual_updated}条')
        except Exception as e:
            db.session.rollback()
            logging.error(f'批量处理失败: {str(e)}')
            raise e
        return jsonify({
            'success': True,
            'message': f'批量处理完成，成功{success_count}条，失败{fail_count}条，实际更新{actual_updated}条',
            'total_count': success_count + fail_count,  # 新增：总记录数
            'success_count': success_count,
            'fail_count': fail_count,
            'actual_updated': actual_updated,
            'errors': errors if errors else None
        })

    except SQLAlchemyError as e:
        db.session.rollback()
        error_msg = f'数据库操作失败: {str(e)}'
        logging.error(error_msg)
        log_operation(
                    user_id=current_user.id,
                    module="utility",
                    operation_type="batch_import_export",
                    action=error_msg,
                    result="失败"
                )
            
        return jsonify({'success': False, 'message': error_msg}), 500
    except Exception as e:
        error_msg = f'处理文件失败: {str(e)}'
        logging.error(error_msg)
        log_operation(
                    user_id=current_user.id,
                    module="utility",
                    operation_type="batch_import_export",
                    action=error_msg,
                    result="失败"
                )
        return jsonify({'success': False, 'message': error_msg}), 500
