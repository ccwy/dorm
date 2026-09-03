from flask import Blueprint, request, jsonify,send_file
from models.utility.utility_room_bill_record import RoomUtilityRecord      #导入主表模型
from models.utility.utility_room_bill_occupant import RoomUtilityOccupant  #导入子表模型
from models.dorm.dorm import Dorm
from models.user.user import User  # 假设存在用户模型
from models.department.department import Department
from utils.db import db
from sqlalchemy.exc import SQLAlchemyError
from utils.log import log_operation
from models.room.room import Room
from datetime import datetime, timedelta  # 修正：移除date，保留datetime和timedelta
import io
import logging  # 确保导入logging模块
from sqlalchemy.exc import SQLAlchemyError
from flask_login import login_required, current_user
from models.fee_subsidy.fee_subsidy_usage import FeeSubsidyUsage  # 导入费用补贴子表
# 导入权限装饰器
from utils.auth import require_permission
# 创建蓝图
utility_room_bill_occupants_bp = Blueprint('utility_room_bill_occupants', __name__, url_prefix='/utility_room_bill_occupants')

# 获取费用明细数据
@utility_room_bill_occupants_bp.route('/api/fee_records', methods=['GET'])
@login_required
@require_permission('utility.view')
def get_fee_records():
    """获取费用明细数据，支持筛选和分页，新增住宿周期信息"""
    try:
        # 获取请求参数
        billing_period = request.args.get('billingPeriod', '')
        building = request.args.get('building', '')
        department = request.args.get('department', '')
        search_keyword = request.args.get('searchInput', '')
        search_type = request.args.get('searchType', '')  # 新增：获取搜索类型
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        
        # 解析账期为日期范围（如2025-04 -> 2025-04-01至2025-04-30，返回datetime类型）
        start_date = end_date = None
        if billing_period and len(billing_period) == 7:  # 格式如YYYY-MM
            year, month = map(int, billing_period.split('-'))
            # 修正：使用datetime而非date
            start_date = datetime(year, month, 1)
            # 获取当月最后一天
            if month == 12:
                end_date = datetime(year, month, 31, 23, 59, 59)  # 增加时间部分
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(days=1)
                end_date = end_date.replace(hour=23, minute=59, second=59)  # 增加时间部分
        
        # 基础查询
        query = db.session.query(
            RoomUtilityRecord,
            RoomUtilityOccupant,
            User.name,  # 查询用户名
            Department.name.label('department'),  # 查询部门信息用于筛选
            Dorm,  # 关联Dorm模型获取住宿日期
            Room  # 选择完整的Room对象，确保可以访问room_number字段
        ).join(
            RoomUtilityOccupant, 
            RoomUtilityRecord.record_id == RoomUtilityOccupant.record_id
        ).join(
            User, 
            RoomUtilityOccupant.user_id == User.id
        ).outerjoin(
            Department, User.department_id == Department.id
        ).join(
            Dorm,  # 关联Dorm表
            db.and_(
                Dorm.user_id == RoomUtilityOccupant.user_id,
                Dorm.room_id == RoomUtilityRecord.room_id
            )
        ).join(
            Room,  # 关联Room表用于获取楼栋信息
            RoomUtilityRecord.room_id == Room.id
        )
        
        # 账期筛选
        if billing_period:
            query = query.filter(RoomUtilityRecord.billing_period == billing_period)
        
        # 楼栋筛选
        if building:
            query = query.filter(Room.building == building)
        
        # 部门筛选 - 确保在所有条件下都能正确应用
        if department:
            query = query.filter(Department.name == department)
        
        # 搜索筛选（根据搜索类型分别处理房间号或姓名）
        if search_keyword:
            # Room表已在基础查询中关联，不需要重复关联
            if search_type == 'room':
                # 房间号搜索，使用精确匹配前模糊匹配后
                query = query.filter(Room.room_number.like(f'%{search_keyword}%'))
            elif search_type == 'name':
                # 姓名搜索
                query = query.filter(User.name.like(f'%{search_keyword}%'))
            else:
                # 默认同时搜索房间号和姓名
                query = query.filter(
                    db.or_(
                        Room.room_number.like(f'%{search_keyword}%'),
                        User.name.like(f'%{search_keyword}%')
                    )
                )
        
        
        # 执行分页查询
        pagination = query.order_by(
            RoomUtilityRecord.billing_period.desc(),
            RoomUtilityRecord.room_id.asc()
        ).paginate(page=page, per_page=per_page)
        
        # 处理查询结果
        records = []
        for item in pagination.items:
            main_record, occupant_record, user_name, user_department, dorm_record, room = item
            
            # 计算账期内的实际住宿周期（datetime类型）
            period_start = start_date if start_date else main_record.start_date
            period_end = end_date if end_date else main_record.end_date
            
            # 确定实际入住和退宿日期（取与账期的交集，datetime比较）
            actual_check_in = max(dorm_record.check_in_date, period_start)
            actual_check_out = dorm_record.check_out_date or period_end
            actual_check_out = min(actual_check_out, period_end)
            
            # 格式化日期时间显示（包含秒）
            check_in_str = actual_check_in.strftime('%Y-%m-%d %H:%M:%S')
            check_out_str = actual_check_out.strftime('%Y-%m-%d %H:%M:%S')
            
            # 获取抄表信息
            electric_reading = f"{main_record.electric_previous} → {main_record.electric_current}" if main_record.electric_previous and main_record.electric_current else ""
            water_reading = f"{main_record.water_previous} → {main_record.water_current}" if main_record.water_previous and main_record.water_current else ""
            
            records.append({
                'billing_period': main_record.billing_period,
                'room_id': main_record.room_id,
                'electric_fee': main_record.billing_electric_fee,
                'water_fee': main_record.billing_water_fee,
                'total_fee': main_record.billing_total_fee,
                'checked_out_total_fee': main_record.checked_out_total_fee,
                'actual_total_fee': main_record.actual_total_fee,
                'room_reduction_fee': main_record.room_reduction_fee,# 新增：房间级减免费用
                'user_name': user_name,
                'user_id': occupant_record.user_id,
                'department': user_department,  # 新增：部门信息
                'person_electric_fee': occupant_record.electric_fee,
                'person_water_fee': occupant_record.water_fee,
                'person_total_fee': occupant_record.total_fee,
                'user_reduction_fee': occupant_record.user_reduction_fee,# 新增：减免费用字段
                'payable_fee': occupant_record.payable_fee,# 新增：用户应付费用字段
                'stay_days': occupant_record.stay_days,
                # 新增：住宿周期信息（包含时间）
                'check_in_date': check_in_str,
                'check_out_date': check_out_str,
                'electric_reading': electric_reading,
                'water_reading': water_reading,
                'record_id': main_record.record_id,
                'occupant_id': occupant_record.id,
                'room_number': room.room_number,  # 新增：房间号字段
                'building': room.building  # 新增：楼栋字段
            })
        
        # 记录成功日志
        log_operation(
            user_id=current_user.id,
            module='utility',
            operation_type='utility_api',
            action=f"查询费用明细 [账期: {billing_period}, 搜索关键词: {search_keyword}, 页码: {page}]",
            result="成功"
        )
        
        # 返回分页数据
        return jsonify({
            'success': True,
            'data': records,
                'room_reduction_fee_field': 'room_reduction_fee',  # 明确指定房间级减免费用字段名
            'pagination': {
                'total': pagination.total,
                'pages': pagination.pages,
                'page': page,
                'per_page': per_page,
                'has_next': pagination.has_next,
                'has_prev': pagination.has_prev
            }
        })
        
    except Exception as e:
        logging.error(f"获取费用记录失败: {str(e)}")
        # 记录失败日志
        log_operation(
            user_id=current_user.id,
            module='utility',
            operation_type='utility_api',
            action=f"查询费用明细失败 [错误: {str(e)}]",
            result="失败"
        )
        return jsonify({'success': False, 'message': f'获取数据失败: {str(e)}'}), 500
    
# 加载账单数据
@utility_room_bill_occupants_bp.route('/api/load_bill', methods=['POST'])
@login_required
@require_permission('utility.view')
def load_bill():
    """加载指定账期的账单数据"""
    try:
        data = request.json
        billing_period = data.get('billingPeriod')
        
        if not billing_period:
            log_operation(
                user_id=current_user.id,
                module='utility',
                operation_type='utility_api',
                action=f"加载账单数据失败 [原因: 未提供账期参数]",
                result="失败"
            )
            return jsonify({'success': False, 'message': '请选择账期'}), 400
            
        # 查找该账期的所有主表记录
        main_records = RoomUtilityRecord.get_by_period(None, period=billing_period)
        
        if not main_records:
            log_operation(
                user_id=current_user.id,
                module='utility',
                operation_type='utility_api',
                action=f"加载账单数据失败 [账期: {billing_period}, 原因: 未找到记录]",
                result="失败"
            )
            return jsonify({
                'success': False, 
                'message': f'未找到{ billing_period }的账单记录'
            }), 404
        
        # 加载对应的子表记录
        record_ids = [r.record_id for r in main_records]
        occupant_records = RoomUtilityOccupant.query.filter(
            RoomUtilityOccupant.record_id.in_(record_ids)
        ).all()
        # 记录成功日志
        log_operation(
            user_id=current_user.id,
            module='utility',
            operation_type='utility_api',
            action=f"加载账单数据 [账期: {billing_period}, 主表记录数: {len(main_records)}, 子表记录数: {len(occupant_records)}]",
            result="成功"
        )
        return jsonify({
            'success': True,
            'message': f'已加载{ billing_period }的账单记录',
            'record_count': len(main_records),
            'occupant_count': len(occupant_records)
        })
        
    except Exception as e:
        logging.error(f"加载账单失败: {str(e)}")
        log_operation(
            user_id=current_user.id,
            module='utility',
            operation_type='utility_api',
            action=f"加载账单数据失败 [错误: {str(e)}]",
            result="失败"
        )
        return jsonify({'success': False, 'message': f'加载失败: {str(e)}'}), 500

# 核算当期账单
@utility_room_bill_occupants_bp.route('/api/calculate_bill', methods=['POST'])
@login_required
@require_permission('utility.calculate')
def calculate_bill():
    """核算指定账期的所有房间费用"""
    try:
        data = request.json
        billing_period = data.get('billingPeriod')
        
        if not billing_period:
            log_operation(
                user_id=current_user.id,
                module='utility',
                operation_type='occupant_fee',
                action=f"核算账单失败 [原因: 未提供账期参数]",
                result="失败"
            )
            return jsonify({'success': False, 'message': '请选择账期'}), 400
            
        # 获取该账期的所有主表记录
        main_records = RoomUtilityRecord.get_by_period(None, period=billing_period)
        
        # 关键修复：按账单开始日期排序，确保补贴按时间顺序使用
        main_records.sort(key=lambda x: x.start_date)

        if not main_records:
            log_operation(
                user_id=current_user.id,
                module='utility',
                operation_type='occupant_fee',
                action=f"核算账单失败 [账期: {billing_period}, 原因: 未找到记录]",
                result="失败"
            )
            return jsonify({
                'success': False, 
                'message': f'未找到{ billing_period }的账单记录，请先创建'
            }), 404
        
        # 关键修复2：初始化全局补贴余额字典，跨房间共享
        global_subsidy_balances = {}

        # 再计算子表分摊
        updated_occupant = 0
        updated_room_count = 0  # 新增：统计处理的房间数量
        for record in main_records:
            occupants, global_subsidy_balances = RoomUtilityOccupant.calculate_room_fee(
                record.record_id,
                user_subsidy_balances=global_subsidy_balances # 核心：共享同一个字典
                )  
            updated_occupant += len(occupants)
            updated_room_count += 1  # 每处理一个主表记录，视为处理一个房间
        
        db.session.commit()
        # 记录成功日志
        log_operation(
            user_id=current_user.id,
            module='utility',
            operation_type='occupant_fee',
            action=f"核算当期账单 [账期: {billing_period}, 房间数: {updated_room_count}, 更新子表记录数: {updated_occupant}]",
            result="成功"
        )
        
        return jsonify({
            'success': True,
            'message': f'{ billing_period }的费用核算完成',
            'updated_room_count': updated_room_count,  # 新增：返回房间数量
            'updated_occupant_count': updated_occupant
        })
        
    except SQLAlchemyError as e:
        db.session.rollback()
        logging.error(f"核算账单数据库错误: {str(e)}")
        log_operation(
            user_id=current_user.id,
            module='utility',
            operation_type='occupant_fee',
            action=f"核算账单失败 [错误: {str(e)}]",
            result="失败"
        )
        return jsonify({'success': False, 'message': f'数据库错误: {str(e)}'}), 500
    except Exception as e:
        db.session.rollback()
        logging.error(f"核算账单失败: {str(e)}")
        log_operation(
            user_id=current_user.id,
            module='utility',
            operation_type='occupant_fee',
            action=f"核算账单失败 [错误: {str(e)}]",
            result="失败"
        )
        return jsonify({'success': False, 'message': f'核算失败: {str(e)}'}), 500

# 删除当期子表账单
@utility_room_bill_occupants_bp.route('/api/clear_current_bill', methods=['POST'])
@login_required
@require_permission('utility.delete')
def clear_current_bill():
    """删除指定账期的子表分摊记录（保留主表数据）"""
    try:
        data = request.json
        billing_period = data.get('billingPeriod')
        
        if not billing_period:
            log_operation(
                user_id=current_user.id,
                module='utility',
                operation_type='delete',
                action=f"删除子表账单失败 [原因: 未提供账期参数]",
                result="失败"
            )
            return jsonify({'success': False, 'message': '请选择账期'}), 400
            
        # 获取该账期的所有主表记录ID
        main_records = RoomUtilityRecord.get_by_period(None, period=billing_period)
        if not main_records:
            log_operation(
                user_id=current_user.id,
                module='utility',
                operation_type='delete',
                action=f"删除子表账单失败 [账期: {billing_period}, 原因: 未找到记录]",
                result="失败"
            )
            return jsonify({
                'success': False, 
                'message': f'未找到{ billing_period }的账单记录'
            }), 404
        
        record_ids = [r.record_id for r in main_records]

        

        # 删除对应子表记录
        deleted_count = RoomUtilityOccupant.query.filter(
            RoomUtilityOccupant.record_id.in_(record_ids)
        ).delete(synchronize_session=False)
        
        period = billing_period
        # 删除对应账期+类型的费用补贴子表记录（双重条件）
        subsidy_usage_deleted = FeeSubsidyUsage.query.filter(
            FeeSubsidyUsage.billing_period == period,
            FeeSubsidyUsage.is_checkout == 3
        ).delete(synchronize_session=False)

        db.session.commit()
        # 记录成功日志
        log_operation(
            user_id=current_user.id,
            module='utility',
            operation_type='delete',
            action=f"删除当期子表账单 [账期: {billing_period}, 删除记录数: {deleted_count}，补贴子表删除数量: {subsidy_usage_deleted}]",
            result="成功"
        )
        return jsonify({
            'success': True,
            'message': f'{ billing_period }的子表账单数据已删除',
            'deleted_count': deleted_count
        })
        
    except SQLAlchemyError as e:
        db.session.rollback()
        logging.error(f"删除账单数据库错误: {str(e)}")
        log_operation(
            user_id=current_user.id,
            module='utility',
            operation_type='delete',
            action=f"删除子表账单失败 [错误: {str(e)}]",
            result="失败"
        )
        return jsonify({'success': False, 'message': f'数据库错误: {str(e)}'}), 500
    except Exception as e:
        db.session.rollback()
        logging.error(f"删除账单失败: {str(e)}")
        log_operation(
            user_id=current_user.id,
            module='utility',
            operation_type='delete',
            action=f"删除子表账单失败 [错误: {str(e)}]",
            result="失败"
        )
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'}), 500

# 删除单条费用记录
@utility_room_bill_occupants_bp.route('/api/delete_fee_record/<int:occupant_id>', methods=['DELETE'])
@login_required
@require_permission('utility.delete')
def delete_fee_record(occupant_id):
    """删除单条人员费用记录"""
    try:
        # 查找记录
        record = RoomUtilityOccupant.query.get(occupant_id)
        if not record:
            log_operation(
                user_id=current_user.id,
                module='utility',
                operation_type='delete',
                action=f"删除单条费用记录失败 [记录ID: {occupant_id}, 原因: 记录不存在]",
                result="失败"
            )
            return jsonify({'success': False, 'message': '记录不存在'}), 404
            
        # 获取关联信息用于返回
        main_record = RoomUtilityRecord.get_by_id(record.record_id)
        user = User.query.get(record.user_id)
        user_name = user.name if user else '未知用户'
        
        # 获取关联的补贴记录条件
        user_id = record.user_id
        room_id = main_record.room_id
        billing_period = main_record.billing_period
        
        # 连带删除关联的费用补贴子表记录（与账期删除保持一致的is_checkout=3条件）
        subsidy_deleted_count = FeeSubsidyUsage.query.filter(
            FeeSubsidyUsage.user_id == user_id,
            FeeSubsidyUsage.room_id == room_id,
            FeeSubsidyUsage.billing_period == billing_period,
            FeeSubsidyUsage.is_checkout == 3
        ).delete(synchronize_session=False)
        
        # 删除主记录
        db.session.delete(record)
        db.session.commit()
        # 根据room_id获取楼栋和房间号
        room = Room.query.get(room_id)
        # 记录成功日志
        log_operation(
            user_id=current_user.id,
            module='utility',
            operation_type='delete',
            action=f"删除单条费用记录 [记录ID: {occupant_id}, 账期: {billing_period}, 房间号: {room.building}{room.room_number},"
                   f"用户: {user_name}], 连带删除补贴记录数量: {subsidy_deleted_count}",
            result="成功"
        )
        
        return jsonify({
            'success': True,
            'message': f'已删除 {billing_period} 账期， {room.building}{room.room_number} 房间， {user_name} 的费用记录',
            'subsidy_deleted_count': subsidy_deleted_count
        })
        
    except SQLAlchemyError as e:
        db.session.rollback()
        logging.error(f"删除单条记录数据库错误: {str(e)}")
        log_operation(
            user_id=current_user.id,
            module='utility',
            operation_type='delete',
            action=f"删除单条费用记录失败 [记录ID: {occupant_id}, 错误: {str(e)}]",
            result="失败"
        )
        return jsonify({'success': False, 'message': f'数据库错误: {str(e)}'}), 500
    except Exception as e:
        db.session.rollback()
        logging.error(f"删除单条记录失败: {str(e)}")
        log_operation(
            user_id=current_user.id,
            module='utility',
            operation_type='delete',
            action=f"删除单条费用记录失败 [记录ID: {occupant_id}, 错误: {str(e)}]",
            result="失败"
        )
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'}), 500


def get_billing_period_dates(billing_period):
    """将账期字符串转换为具体日期时间范围（YYYY-MM -> 月初和月末，精确到秒）"""
    try:
        year, month = map(int, billing_period.split('-'))
        # 修正：返回datetime类型，包含时间信息
        start_date = datetime(year, month, 1, 0, 0, 0)  # 月初00:00:00
        # 计算月末日期时间
        if month == 12:
            next_month = 1
            next_year = year + 1
        else:
            next_month = month + 1
            next_year = year
        # 月末最后一秒
        end_date = datetime(next_year, next_month, 1, 0, 0, 0) - timedelta(seconds=1)
        return start_date, end_date
    except Exception as e:
        # 修复日志记录方式 - 使用logging并包含详细信息
        logging.error(f"解析账期失败: {billing_period}, 错误: {str(e)}")
        raise ValueError(f"无效的账期格式: {billing_period}，应为YYYY-MM")

def create_fee_export_data(billing_period):
    """创建导出数据，仅按账期筛选"""
    # 构建查询条件 - 仅按账期筛选
    query = RoomUtilityRecord.query.filter(RoomUtilityRecord.billing_period == billing_period)
    
    # 预计算账期日期范围用于后续天数计算（datetime类型）
    start_date, end_date = get_billing_period_dates(billing_period)
    
    # 获取符合条件的主表记录
    main_records = query.all()
    export_data = []
    
    for main in main_records:
        # 获取该记录的所有人员分摊记录
        occupant_records = RoomUtilityOccupant.get_by_record(main.record_id)
        
        # 获取房间信息
        room = Room.query.get(main.room_id)
        if not room:
            logging.error(f"找不到房间信息: room_id={main.room_id}")
            # 跳过此记录，避免后续错误
            continue
        
        # 验证房间信息完整性
        if not room.building or not room.room_number:
            logging.error(f"房间信息不完整: room_id={main.room_id}, building={room.building}, room_number={room.room_number}")
            # 跳过此记录
            continue
        
        # 批量获取住宿记录（包括换宿历史）
        user_ids = [rec.user_id for rec in occupant_records]
        # 获取所有状态的住宿记录，不仅仅是active
        all_dorm_records = Dorm.query.filter(
            Dorm.user_id.in_(user_ids),
            Dorm.room_id == main.room_id
        ).all()
        
        # 构建用户住宿记录映射，考虑换宿链
        dorm_map = {}
        for dorm in all_dorm_records:
            # 对于每个用户，获取其完整住宿链
            user_dorms = dorm.dorm_chain
            # 在账期内有效的住宿记录
            valid_dorms = []
            for d in user_dorms:
                # 检查住宿记录是否在账期内
                dorm_end_date = d.check_out_date if d.check_out_date else end_date
                if not (d.check_in_date > end_date or dorm_end_date < start_date):
                    valid_dorms.append(d)
            
            # 按入住日期排序，取最新的有效记录
            if valid_dorms:
                valid_dorms.sort(key=lambda x: x.check_in_date, reverse=True)
                dorm_map[dorm.user_id] = valid_dorms[0]
        
        # 批量获取用户信息
        users = User.query.filter(User.id.in_(user_ids)).all()
        user_map = {user.id: user for user in users}
        
        # 收集数据
        for occupant in occupant_records:
            user = user_map.get(occupant.user_id)
            user_name = user.name if user else f"未知用户（ID:{occupant.user_id}）"
            user_company = user.company or "" if user else ""
            user_department = user.department or "" if user else ""
            user_position = user.position or "" if user else ""
            
            dorm = dorm_map.get(occupant.user_id)
            check_in_date = dorm.check_in_date if dorm else None
            # 格式化日期时间显示
            check_in_str = check_in_date.strftime('%Y-%m-%d') if check_in_date else ""
            
            # 构建导出记录
            export_data.append({
                '账期': main.billing_period,
                '房间ID': main.room_id,
                '楼栋': room.building,
                '房间号': room.room_number,
                '本期电表当前读数': main.electric_current,
                '本期电表上期读数': main.electric_previous,
                '本期电表用量': main.electric_usage,
                '减免电用量': main.electric_reduction,
                '计费电用量': main.electric_billing_usage,
                '电费单价': main.electric_price,
                '本期电费': main.total_electric_fee,
                '计费电费': main.billing_electric_fee,
                '本期水表当前读数': main.water_current,
                '本期水表上期读数': main.water_previous,
                '本期水表用量': main.water_usage,
                '减免水用量': main.water_reduction,
                '计费水用量': main.water_billing_usage,
                '水费单价': main.water_price,
                '本期水费': main.total_water_fee,
                '计费水费': main.billing_water_fee,
                '本期总费用': main.total_fee,
                '计费总费用': main.billing_total_fee,
                '退宿人员费用': main.checked_out_total_fee,
                '减免房间级费用': main.room_reduction_fee,
                '房间应付费用': main.actual_total_fee,
                '分摊人员ID': occupant.user_id,
                '分摊人员姓名': user_name,
                '公司': user_company,
                '部门': user_department,
                '职位': user_position,
                '入住时间': check_in_str,  # 已转换为包含时间的字符串
                '账期内住宿天数': occupant.stay_days,
                '分摊电费': occupant.electric_fee,
                '分摊水费': occupant.water_fee,
                '分摊总金额': occupant.total_fee,
                '减免金额': occupant.user_reduction_fee, # 新增：减免费用字段
                '分摊应付金额': occupant.payable_fee # 新增：用户应付费用字段
            })
    
    return export_data

@utility_room_bill_occupants_bp.route('/api/export_fee_data', methods=['GET'])
@login_required
@require_permission('utility.export')
def export_fee_data():
    """导出人员费用数据为Excel，支持按房间号合并所有相同内容字段并添加完整边框"""
    import pandas as pd  # 延迟导入，避免启动时加载重型库
    try:
        # 获取筛选参数
        billing_period = request.args.get('billing_period') or request.args.get('billingPeriod')
        
        # 验证账期参数
        if not billing_period:
            # 修复日志记录 - 使用log_operation函数记录操作结果
            log_operation(
                user_id=current_user.id,
                module="utility",
                operation_type="batch_import_export",
                action=f"导出费用数据失败 [原因: 未提供账期参数]",
                result="失败"
            )
            # 同时记录到logging
            logging.warning(f"用户 {current_user.id} 未提供账期参数尝试导出费用数据")
            return jsonify({
                'success': False,
                'message': '请提供账期参数(billing_period，格式为YYYY-MM)'
            }), 400
        
        # 记录导出操作开始 - 与日志蓝图保持一致的记录方式
        log_operation(
            user_id=current_user.id,
            module="utility",
            operation_type="batch_import_export",
            action=f"开始导出费用数据 [账期: {billing_period}]",
            result="开始"
        )
        logging.info(f"用户 {current_user.id} 开始导出 {billing_period} 账期的费用数据")
        
        # 创建导出数据
        export_data = create_fee_export_data(billing_period)
        
        if not export_data:
            # 记录无数据情况
            log_operation(
                user_id=current_user.id,
                module="utility",
                operation_type="batch_import_export",
                action=f"导出费用数据失败 [账期: {billing_period}, 原因: 未找到匹配记录]",
                result="失败"
            )
            logging.info(f"用户 {current_user.id} 导出 {billing_period} 账期费用数据，未找到匹配记录")
            return jsonify({
                'success': False,
                'message': f'没有找到{ billing_period }账期的费用数据'
            }), 404
        
        # 创建Excel
        df = pd.DataFrame(export_data)
        
        # 先按楼栋排序，再按房间号排序，确保不同楼栋的相同房间号不会被混淆
        df = df.sort_values(by=['楼栋', '房间号'])
        
        # 处理日期时间格式（保留时间信息）
        if '入住时间' in df.columns:
            # 转换为datetime类型保留完整信息
            df['入住时间'] = pd.to_datetime(df['入住时间'], format='%Y-%m-%d', errors='coerce').dt.date
        
        # 保存到内存
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='人员费用分摊')
            
            # 获取工作表对象
            worksheet = writer.sheets['人员费用分摊']
            
            # 定义样式
            from openpyxl.styles import Border, Side, Alignment
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # 所有单元格统一使用垂直居中，表头额外使用水平居中
            data_alignment = Alignment(vertical='center')
            header_alignment = Alignment(vertical='center', horizontal='center')
            
            # 需要合并的列名（房间相关的公共信息）
            columns_to_merge = [
                '账期', '房间ID', '楼栋', '房间号',
                '本期电表当前读数', '本期电表上期读数', '本期电表用量', '减免电用量', '计费电用量', '电费单价','本期电费', '计费电费',
                '本期水表当前读数', '本期水表上期读数', '本期水表用量', '减免水用量', '计费水用量', '水费单价','本期水费', '计费水费',
                '本期总费用', '计费总费用', '退宿人员费用', '减免房间级费用', '房间应付费用'
            ]
            
            # 存储列名到索引的映射（1-based）
            col_index_map = {}
            for col_idx, cell in enumerate(worksheet[1]):  # 表头行
                col_index_map[cell.value] = col_idx + 1  # openpyxl是1-based索引
                # 设置表头样式
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 获取所有数据行
            max_row = worksheet.max_row
            max_col = len(col_index_map)
            
            # 先为所有单元格应用基础样式（边框和垂直居中）
            for row in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = data_alignment
                    cell.border = thin_border
            
            if max_row > 1:  # 确保有数据行
                # 获取楼栋和房间号列的索引
                building_col_idx = col_index_map.get('楼栋')
                room_col_idx = col_index_map.get('房间号')
                
                if building_col_idx and room_col_idx:
                    # 记录当前楼栋、房间号和起始行
                    current_building = worksheet.cell(row=2, column=building_col_idx).value
                    current_room = worksheet.cell(row=2, column=room_col_idx).value
                    start_row = 2
                    
                    # 遍历所有行，识别连续相同的楼栋和房间号组合
                    for row in range(3, max_row + 1):
                        building_value = worksheet.cell(row=row, column=building_col_idx).value
                        room_value = worksheet.cell(row=row, column=room_col_idx).value
                        
                        if building_value != current_building or room_value != current_room:
                            # 对所有需要合并的列执行合并操作
                            for col_name in columns_to_merge:
                                col_idx = col_index_map.get(col_name)
                                if col_idx and (row - start_row > 1):
                                    # 合并单元格
                                    worksheet.merge_cells(
                                        start_row=start_row, 
                                        start_column=col_idx,
                                        end_row=row - 1, 
                                        end_column=col_idx
                                    )
                                    # 合并后重新设置样式（合并会清除部分样式）
                                    merged_cell = worksheet.cell(row=start_row, column=col_idx)
                                    merged_cell.alignment = Alignment(vertical='center', horizontal='center')
                                    merged_cell.border = thin_border
                            
                            current_building = building_value
                            current_room = room_value
                            start_row = row
                    
                    # 处理最后一组相同楼栋和房间号的行
                    if max_row - start_row > 0:
                        for col_name in columns_to_merge:
                            col_idx = col_index_map.get(col_name)
                            if col_idx:
                                worksheet.merge_cells(
                                    start_row=start_row, 
                                    start_column=col_idx,
                                    end_row=max_row, 
                                    end_column=col_idx
                                )
                                # 合并后重新设置样式
                                merged_cell = worksheet.cell(row=start_row, column=col_idx)
                                merged_cell.alignment = Alignment(vertical='center', horizontal='center')
                                merged_cell.border = thin_border
        
        output.seek(0)
        
        # 构建文件名
        filename = f"人员费用分摊数据_{billing_period}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        logging.info(f"用户 {current_user.id} 导出 {billing_period} 账期费用数据，文件名: {filename}")
        # 记录导出成功
        log_operation(
            user_id=current_user.id,
            module="utility",
            operation_type="batch_import_export",
            action=f"导出费用数据成功 [账期: {billing_period}, 记录数: {len(export_data)}]",
            result="成功"
        )
        logging.info(f"用户 {current_user.id} 成功导出 {billing_period} 账期费用数据，共 {len(export_data)} 条记录")
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
        
    except ValueError as ve:
        # 参数错误日志记录
        logging.warning(f"用户 {current_user.id} 导出费用数据参数错误: {str(ve)}")
        log_operation(
            user_id=current_user.id,
            module="utility",
            operation_type="batch_import_export",
            action=f"导出费用数据失败 [错误: {str(ve)}]",
            result="失败"
        )
        return jsonify({
            'success': False,
            'message': str(ve)
        }), 400
    except Exception as e:
        # 异常错误日志记录
        logging.error(f"用户 {current_user.id} 导出费用数据失败: {str(e)}", exc_info=True)
        log_operation(
            user_id=current_user.id,
            module="utility",
            operation_type="batch_import_export",
            action=f"导出费用数据失败 [错误: {str(e)}]",
            result="失败"
        )
        return jsonify({
            'success': False,
            'message': f'导出失败: {str(e)}'
        }), 500
